import os
from fastapi import FastAPI, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from transformers import CLIPProcessor, CLIPModel
from PIL import Image
import torch
import numpy as np
import faiss
from collections import defaultdict
import io
from openpyxl import load_workbook

from fastapi.responses import HTMLResponse
from fastapi.templating import Jinja2Templates
from fastapi import Request

app = FastAPI(title="Fashion Image Search")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ============= CONFIG =============
MODEL_NAME = "patrickjohncyh/fashion-clip"
EMBED_DIM = 512
MEDIA_DIR = "media/products"
os.makedirs(MEDIA_DIR, exist_ok=True)

W_GLOBAL, W_CLOTHES, W_RERANK = 0.2, 0.5, 0.3
SEARCH_K_FACTOR = 3
# ==================================

# Serve ảnh tĩnh
app.mount("/media", StaticFiles(directory="media"), name="media")
app.mount("/static", StaticFiles(directory="static"), name="static")

templates = Jinja2Templates(directory="templates")

@app.get("/", response_class=HTMLResponse)
async def home(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})

device = "cuda" if torch.cuda.is_available() else "cpu"
print(f"Loading {MODEL_NAME} on {device}")
model = CLIPModel.from_pretrained(MODEL_NAME).to(device)
processor = CLIPProcessor.from_pretrained(MODEL_NAME, use_fast=True)
model.eval()

# FAISS index
index_global = faiss.IndexFlatIP(EMBED_DIM)
index_clothes = faiss.IndexFlatIP(EMBED_DIM)

product_ids_global, product_ids_clothes = [], []
embeddings_global_store, embeddings_clothes_store = [], []

# Lưu thông tin sản phẩm
products = {}  # {code: {"images":[], "notes":[]}}

# ====== Hàm embedding ======
def get_image_embedding(image: Image.Image):
    inputs = processor(images=image, return_tensors="pt").to(device)
    with torch.no_grad():
        emb = model.get_image_features(**inputs)
    emb = emb / emb.norm(p=2, dim=-1, keepdim=True)
    return emb.cpu().numpy()

# ====== Hàm crop áo ======
def detect_clothes(image: Image.Image) -> Image.Image:
    w, h = image.size
    left, top = int(w * 0.15), int(h * 0.15)
    right, bottom = int(w * 0.85), int(h * 0.85)
    return image.crop((left, top, right, bottom))

def chunks(lst, batch_size):
    """Yield successive batch_size-sized chunks from lst."""
    for i in range(0, len(lst), batch_size):
        yield lst[i:i + batch_size]

@app.post("/add-excel/")
async def add_from_excel(file: UploadFile = File(...), batch_size: int = 50):
    # Stream file thay vì đọc toàn bộ vào bộ nhớ
    temp_file_path = f"temp_excel_{id(file)}.xlsx"
    wb = None
    try:
        # Lưu file tạm thời để streaming
        with open(temp_file_path, "wb") as f:
            # Đọc và ghi từng chunk nhỏ
            chunk_size = 1024 * 1024  # 1MB chunks
            while chunk := await file.read(chunk_size):
                f.write(chunk)
        
        # Đọc workbook từ file tạm - BỎ read_only=True
        wb = load_workbook(temp_file_path)  # Bỏ read_only để có thể truy cập _images
        sheet = wb.active
        
        if not hasattr(sheet, '_images') or not sheet._images:
            return {"status": "error", "message": "Không tìm thấy hình ảnh trong file Excel"}
        
        # Xử lý hình ảnh theo batch nhỏ
        img_count = len(sheet._images)
        added = []
        
        # Phần còn lại của hàm giữ nguyên
        for i in range(0, img_count, batch_size):
            batch_images = sheet._images[i:i+batch_size]
            codes, notes = [], []
            images, clothes_images = [], []
            
            # Xử lý từng hình ảnh trong batch
            for img_idx, img in enumerate(batch_images):
                try:
                    row = img.anchor._from.row + 1
                    code_val = sheet.cell(row, 2).value  # B = mã
                    note_val = sheet.cell(row, 3).value  # C = ghi chú
                    
                    if not code_val:
                        continue
                    
                    code = str(code_val).strip()
                    note = str(note_val).strip() if note_val else ""
                    
                    # Chỉ xử lý từng hình ảnh một
                    pil_img = Image.open(io.BytesIO(img._data())).convert("RGB")
                    
                    # Lưu hình ảnh ngay lập tức
                    filename = f"{code}_{len(products.get(code, {}).get('images', [])) + 1}.jpg"
                    filepath = os.path.join(MEDIA_DIR, filename)
                    pil_img.save(filepath)
                    url = f"/media/products/{filename}"
                    
                    # Cập nhật từ điển product
                    if code not in products:
                        products[code] = {"images": [], "notes": []}
                    products[code]["images"].append(url)
                    if note:
                        products[code]["notes"].append(note)
                    
                    # Thêm vào batch hiện tại
                    codes.append(code)
                    notes.append(note)
                    images.append(pil_img)
                    clothes_images.append(detect_clothes(pil_img))
                    
                except Exception as e:
                    print(f"Lỗi xử lý ảnh {i + img_idx + 1}: {str(e)}")
            
            # Xử lý embedding cho batch hiện tại
            if not images:
                continue  # Bỏ qua nếu không có hình ảnh nào
                
            try:
                # Xử lý toàn cảnh
                inputs_global = processor(images=images, return_tensors="pt", padding=True).to(device)
                with torch.no_grad():
                    embs_global = model.get_image_features(**inputs_global)
                embs_global = embs_global / embs_global.norm(p=2, dim=-1, keepdim=True)
                embs_global = embs_global.cpu().numpy()
                
                # Giải phóng bộ nhớ ngay lập tức
                inputs_global = inputs_global.to("cpu")
                del inputs_global
                torch.cuda.empty_cache()
                
                # Xử lý hình ảnh quần áo
                inputs_clothes = processor(images=clothes_images, return_tensors="pt", padding=True).to(device)
                with torch.no_grad():
                    embs_clothes = model.get_image_features(**inputs_clothes)
                embs_clothes = embs_clothes / embs_clothes.norm(p=2, dim=-1, keepdim=True)
                embs_clothes = embs_clothes.cpu().numpy()
                
                # Giải phóng bộ nhớ ngay lập tức
                inputs_clothes = inputs_clothes.to("cpu")
                del inputs_clothes
                torch.cuda.empty_cache()
                
                # Thêm vào index
                for idx, (code, note, g_emb, c_emb) in enumerate(zip(codes, notes, embs_global, embs_clothes)):
                    index_global.add(g_emb.reshape(1, -1))
                    product_ids_global.append(code)
                    embeddings_global_store.append(g_emb)
                    
                    index_clothes.add(c_emb.reshape(1, -1))
                    product_ids_clothes.append(code)
                    embeddings_clothes_store.append(c_emb)
                    
                    added.append({"product_id": code, "note": note})
                
            finally:
                # Đảm bảo giải phóng bộ nhớ
                for img in images:
                    del img
                for img in clothes_images:
                    del img
                del images[:]
                del clothes_images[:]
                del embs_global
                del embs_clothes
                torch.cuda.empty_cache()
                
                # Gọi garbage collector
                import gc
                gc.collect()
        
        return {"status": "ok", "added_count": len(added)}
        
    except Exception as e:
        return {"status": "error", "message": str(e)}
    finally:
        # Đóng workbook trước khi xóa file tạm
        if wb is not None:
            wb.close()
            # Đảm bảo workbook được giải phóng hoàn toàn
            del wb
            import gc
            gc.collect()
            
        # Thêm một chút delay để đảm bảo file đã được giải phóng
        import time
        time.sleep(0.5)
        
        # Thử xóa file với retry
        max_retries = 5
        for i in range(max_retries):
            try:
                if os.path.exists(temp_file_path):
                    os.remove(temp_file_path)
                break
            except PermissionError:
                if i < max_retries - 1:
                    time.sleep(0.5)  # Đợi thêm 0.5 giây trước khi thử lại
                else:
                    print(f"Không thể xóa file tạm: {temp_file_path}. File sẽ được giữ lại.")

@app.post("/search/")
async def search(file: UploadFile = File(...), top_k: int = 5):
    img_bytes = await file.read()
    image = Image.open(io.BytesIO(img_bytes)).convert("RGB")

    q_global = get_image_embedding(image)[0]
    q_clothes = get_image_embedding(detect_clothes(image))[0]

    k_pull = max(top_k * SEARCH_K_FACTOR, top_k)
    candidate_scores = defaultdict(list)

    if index_global.ntotal > 0:
        s, idxs = index_global.search(q_global.reshape(1, -1), min(k_pull, index_global.ntotal))
        for score, idx in zip(s[0], idxs[0]):
            pid = product_ids_global[idx]
            candidate_scores[pid].append(("global", float(score)))

    if index_clothes.ntotal > 0:
        s, idxs = index_clothes.search(q_clothes.reshape(1, -1), min(k_pull, index_clothes.ntotal))
        for score, idx in zip(s[0], idxs[0]):
            pid = product_ids_clothes[idx]
            candidate_scores[pid].append(("clothes", float(score)))

    final_list = []
    for pid, entries in candidate_scores.items():
        best_global = max((s for src, s in entries if src == "global"), default=0.0)
        best_clothes = max((s for src, s in entries if src == "clothes"), default=0.0)
        rerank_score = max(
            (float(np.dot(q_clothes, emb)) for emb, p in zip(embeddings_clothes_store, product_ids_clothes) if p == pid),
            default=0.0
        )
        final_score = W_GLOBAL*best_global + W_CLOTHES*best_clothes + W_RERANK*rerank_score

        final_list.append({
            "product_id": pid,
            "score": final_score,
            "images": products[pid]["images"],
            "notes": products[pid]["notes"]
        })

    final_list = sorted(final_list, key=lambda x: x["score"], reverse=True)[:top_k]
    return {"results": final_list}

@app.get("/health")
async def health():
    return {"status": "ok", "model": MODEL_NAME, "device": device}

if __name__ == "__main__":
    import os
    import uvicorn

    port = int(os.environ.get("PORT", 8000))  # Railway sẽ inject PORT
    uvicorn.run("qdrant_image3:app", host="0.0.0.0", port=port)