from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
import fitz  # PyMuPDF
import os

# === CẤU HÌNH ===
PDF_FILE = "file.pdf"  # ← Đổi tên file PDF
EXCEL_FILE = "file.xlsx"  # ← Đổi tên file Excel chứa mã + ghi chú
OUTPUT_EXCEL = "export_with_images.xlsx"
OUTPUT_IMG_DIR = "extracted_images"

# === TRÍCH ẢNH TỪ PDF ===
os.makedirs(OUTPUT_IMG_DIR, exist_ok=True)
doc = fitz.open(PDF_FILE)
img_paths = []

for page_number in range(len(doc)):
    page = doc[page_number]
    images = page.get_images(full=True)
    for img_index, img in enumerate(images):
        xref = img[0]
        base_image = doc.extract_image(xref)
        image_bytes = base_image["image"]
        image_ext = base_image["ext"]
        image_filename = f"page{page_number+1}_img{img_index+1}.{image_ext}"
        image_path = os.path.join(OUTPUT_IMG_DIR, image_filename)
        with open(image_path, "wb") as f:
            f.write(image_bytes)
        img_paths.append(image_path)

# === ĐỌC EXCEL VÀ GHÉP ẢNH ===
wb = load_workbook(EXCEL_FILE, data_only=True)
ws = wb.active

data_rows = []
for row in ws.iter_rows(min_row=2, values_only=True):  # Bỏ dòng tiêu đề
    code = str(row[0]).strip() if row[0] else ""
    note = str(row[1]).strip() if row[1] else ""
    data_rows.append((code, note))

# === GHÉP ẢNH VÀO EXCEL MỚI ===
export_wb = Workbook()
export_ws = export_wb.active
export_ws.title = "Danh sách ảnh"
export_ws.append(["Ảnh", "Mã sản phẩm", "Ghi chú"])

for i, (code, note) in enumerate(data_rows):
    row_excel = i + 2
    if i < len(img_paths):
        img_path = img_paths[i]
        xl_img = XLImage(img_path)
        xl_img.width = 100
        xl_img.height = 100
        export_ws.row_dimensions[row_excel].height = 80
        export_ws.add_image(xl_img, f"A{row_excel}")
    export_ws.cell(row=row_excel, column=2, value=code)
    export_ws.cell(row=row_excel, column=3, value=note)

export_wb.save(OUTPUT_EXCEL)
print(f"✅ Đã xuất file: {OUTPUT_EXCEL}")
