from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage
import io, os

def get_cell_position_from_anchor(anchor):
    if hasattr(anchor, '_from'):
        cell = anchor._from
        row = cell.row + 1
        col = cell.col + 1
        col_letter = get_column_letter(col)
        cell_ref = f"{col_letter}{row}"
        return row, col, cell_ref
    else:
        return None, None, None

def extract_images_with_codes(input_excel, output_excel="hd-out-set-thun.xlsx", image_dir="output_images"):
    os.makedirs(image_dir, exist_ok=True)

    wb = load_workbook(input_excel, data_only=True)
    export_wb = Workbook()
    export_ws = export_wb.active
    export_ws.title = "Danh sách ảnh"

    export_ws.append(["Ảnh", "Mã sản phẩm", "Ghi chú"])

    row_output = 2

    for sheet in wb.worksheets:
        if not hasattr(sheet, "_images"):
            continue

        last_product_code = None  # Lưu mã sản phẩm gần nhất (nếu gặp None thì dùng mã này)

        for idx, img in enumerate(sheet._images):
            if not isinstance(img, XLImage):
                continue

            anchor = img.anchor
            row, col, cell_ref = get_cell_position_from_anchor(anchor)

            # Mã sản phẩm ở dòng dưới ảnh (col giữ nguyên)
            code_row = row + 1
            code_col_letter = get_column_letter(col)
            code_cell = f"{code_col_letter}{code_row}"
            product_code = sheet[code_cell].value

            # Nếu product_code None thì lấy lại mã trước đó
            if product_code is None:
                product_code = last_product_code
            else:
                last_product_code = product_code

            # Ghi chú lấy từ 2 ô tiếp theo bên dưới (vd A3 & A4 nếu ảnh ở A1)
            note_row_1 = row + 2
            note_row_2 = row + 3
            note_cell_1 = f"{code_col_letter}{note_row_1}"
            note_cell_2 = f"{code_col_letter}{note_row_2}"

            note_1 = sheet[note_cell_1].value
            note_2 = sheet[note_cell_2].value

            # Nối nội dung ghi chú (bỏ None)
            notes = []
            if note_1 is not None and str(note_1).strip() != "":
                notes.append(str(note_1).strip())
            if note_2 is not None and str(note_2).strip() != "":
                notes.append(str(note_2).strip())
            ghi_chu = " | ".join(notes) if notes else ""

            # Lưu ảnh ra file
            if isinstance(img.ref, io.BytesIO):
                img_bytes = img.ref
                img_bytes.seek(0)
                pil_img = PILImage.open(img_bytes)
            else:
                pil_img = PILImage.open(img.ref)

            img_path = os.path.join(image_dir, f"{sheet.title}_{cell_ref}.png")
            pil_img.save(img_path, format="PNG")

            max_height = 100  # bạn có thể đổi tùy ý
            w, h = pil_img.size

            if h > max_height:
                ratio = max_height / h
                new_w = int(w * ratio)
                new_h = int(h * ratio)
            else:
                new_w = w
                new_h = h

            export_img = XLImage(img_path)
            export_img.width = new_w
            export_img.height = new_h

            export_ws.row_dimensions[row_output].height = 80  # đặt chiều cao dòng cho vừa ảnh
            export_ws.add_image(export_img, f"A{row_output}")

            export_ws.cell(row=row_output, column=2, value=str(product_code))
            export_ws.cell(row=row_output, column=3, value=ghi_chu)

            row_output += 1  # chừa chỗ cho ảnh

    export_wb.save(output_excel)
    print(f"✅ Đã xuất file: {output_excel}")

# === CHẠY ===
if __name__ == "__main__":
    input_file = "hd-set-thun.xlsx"  # ← Đổi tên file gốc ở đây
    extract_images_with_codes(input_file)
